import os
from flask import Flask, render_template, request, jsonify, send_from_directory
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

app = Flask(__name__)

PROCESSED_FOLDER = 'processed'
os.makedirs(PROCESSED_FOLDER, exist_ok=True)

REQUIRED_COLUMNS = [
    "Recipient's GSTIN", "Type of supply", "Taxable Value", "IGST", "CGST", "SGST/UTGST"
]
NUMERIC_COLUMNS = ["Taxable Value", "IGST", "CGST", "SGST/UTGST"]

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file part'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'No file selected'}), 400

    try:
        df = pd.read_excel(file, skiprows=6)
        df.columns = df.columns.str.strip()

        column_mapping = {
            'Recipients GSTIN': "Recipient's GSTIN",
            'Type of supply': 'Type of supply',
            'Taxable Value (Rs.)': 'Taxable Value',
            'IGST \n(Rs.)': 'IGST',
            'CGST \n(Rs.)': 'CGST',
            'SGST/UTGST \n(Rs.)': 'SGST/UTGST'
        }
        df.rename(columns=column_mapping, inplace=True)

        missing_cols = [col for col in REQUIRED_COLUMNS if col not in df.columns]
        if missing_cols:
            return jsonify({'success': False, 'error': f'Missing required columns: {", ".join(missing_cols)}'}), 400

        df = df[REQUIRED_COLUMNS].dropna(subset=["Recipient's GSTIN"])
        df = df[df["Recipient's GSTIN"].astype(str).str.len() == 15]

        for col in NUMERIC_COLUMNS:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

        summary = df.groupby('Type of supply')[NUMERIC_COLUMNS].sum().reset_index()

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'processed_summary_{timestamp}.xlsx'
        output_path = os.path.join(PROCESSED_FOLDER, filename)

        create_formatted_excel(summary, output_path)

        return jsonify({'success': True, 'download_url': f'/download/{filename}', 'filename': filename})

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

def create_formatted_excel(df, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Processed Summary"
    
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    data_font = Font(size=11)
    data_alignment = Alignment(horizontal="left", vertical="center")
    number_alignment = Alignment(horizontal="right", vertical="center")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for i, cell in enumerate(row):
            cell.border = border
            cell.font = data_font
            
            if i == 0:
                cell.alignment = data_alignment
            else:
                cell.alignment = number_alignment
                if cell.value is not None:
                    try:
                        cell.value = float(cell.value)
                        cell.number_format = '#,##0.00'
                    except:
                        pass
    
    column_widths = {
        'A': 25,
        'B': 15,
        'C': 12,
        'D': 12,
        'E': 15
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    ws.freeze_panes = "A2"
    
    add_summary_statistics(wb, df)
    
    wb.save(output_path)

def add_summary_statistics(wb, df):
    ws_stats = wb.create_sheet("Summary Statistics")
    
    total_taxable = df['Taxable Value'].sum()
    total_igst = df['IGST'].sum()
    total_cgst = df['CGST'].sum()
    total_sgst = df['SGST/UTGST'].sum()
    total_tax = total_igst + total_cgst + total_sgst
    num_categories = len(df)
    
    stats_data = [
        ["Summary Statistics", ""],
        ["", ""],
        ["Total Categories", num_categories],
        ["Total Taxable Value", total_taxable],
        ["Total IGST", total_igst],
        ["Total CGST", total_cgst],
        ["Total SGST/UTGST", total_sgst],
        ["Total Tax Amount", total_tax],
        ["", ""],
        ["Category Breakdown", ""],
    ]
    
    for _, row in df.iterrows():
        stats_data.append([row['Type of supply'], row['Taxable Value']])
    
    for row_data in stats_data:
        ws_stats.append(row_data)
    
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    data_font = Font(size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in ws_stats[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    for row in ws_stats.iter_rows(min_row=2, max_row=ws_stats.max_row):
        for i, cell in enumerate(row):
            cell.border = border
            cell.font = data_font
            
            if i == 1 and cell.value is not None and isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
    
    ws_stats.column_dimensions['A'].width = 25
    ws_stats.column_dimensions['B'].width = 20

@app.route('/download/<filename>')
def download(filename):
    return send_from_directory(PROCESSED_FOLDER, filename, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)
